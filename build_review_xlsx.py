import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

data = json.load(open('data/messy-records.json', encoding='utf-8'))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "На перевірку"

headers = ["Назва", "Категорія", "Проблема", "Адреса (розпізнано)", "Телефон (розпізнано)",
           "Адреса (виправлено)", "Телефон (виправлено)", "Опис", "Сирий текст джерела (довідково)"]

header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

notes = {
    "Назва": "Не редагувати, тільки звіряти",
    "Проблема": "Що саме виглядає підозріло: адреса, телефон, або обидва",
    "Адреса (виправлено)": "Впиши сюди правильну адресу, звірившись із сирим текстом справа",
    "Телефон (виправлено)": "Впиши сюди правильний телефон у форматі +380 XX XXX XX XX",
    "Сирий текст джерела (довідково)": "Оригінальний нерозібраний текст — звідки бралось усе. Може містити правильні цифри/адресу, яких не вдалось розпізнати автоматично",
}
for col, h in enumerate(headers, start=1):
    if h in notes:
        ws.cell(row=1, column=col).comment = Comment(notes[h], "M-ReinWerk")

fill_input = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")

for i, rec in enumerate(data, start=2):
    row = [
        rec['name'], rec['category'], rec['issue'],
        rec['address_parsed'], rec['phone_parsed'],
        "", "",
        rec['description'], rec['raw_blob']
    ]
    for col, val in enumerate(row, start=1):
        c = ws.cell(row=i, column=col, value=val)
        c.font = Font(name="Arial", size=10)
        c.border = border
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if col in (6, 7):
            c.fill = fill_input

widths = [22, 22, 14, 26, 20, 26, 20, 30, 55]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 34

wb.save('/mnt/user-data/outputs/starokostiantyniv_na_perevirku.xlsx')
print('Збережено, рядків:', len(data))
