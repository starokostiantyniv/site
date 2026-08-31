import json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

data = json.load(open('data/all-records-v2.json', encoding='utf-8'))
flagged = [r for r in data if r['flags']]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "На перевірку"

headers = ["Назва (розпізнано)", "Категорія", "Проблема", "Рекомендація",
           "Адреса (розпізнано)", "Телефон (розпізнано)",
           "Назва (виправлено)", "Адреса (виправлено)", "Телефон (виправлено)",
           "Опис", "Сирий текст джерела (довідково)"]

header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

notes = {
    "Проблема": "Може бути кілька через кому: адреса / телефон / можливо це нік, а не назва / обрізана назва / артефакт (не заклад)",
    "Рекомендація": "Автоматична підказка. Остаточне рішення — за тобою",
    "Назва (виправлено)": "Заповнюй, тільки якщо назва підозріла (нік/обрізана/артефакт). Якщо назва ок — лиши пустим",
    "Адреса (виправлено)": "Заповнюй, тільки якщо є позначка 'адреса'",
    "Телефон (виправлено)": "Заповнюй, тільки якщо є позначка 'телефон'",
    "Сирий текст джерела (довідково)": "Оригінал, звідки все розбиралось. Порожньо для рядків з Аркуш1 (там і так чисто)",
}
for col, h in enumerate(headers, start=1):
    if h in notes:
        ws.cell(row=1, column=col).comment = Comment(notes[h], "M-ReinWerk")

fill_input = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
fill_delete = PatternFill(start_color="FECACA", end_color="FECACA", fill_type="solid")

def recommendation(flags):
    if 'артефакт (не заклад)' in flags:
        return "Видалити рядок — це не заклад"
    if 'можливо це нік, а не назва' in flags and 'обрізана назва' in flags:
        return "Перевірити назву і адресу"
    if 'можливо це нік, а не назва' in flags:
        return "Перевірити, чи це справжня назва бізнесу"
    if 'обрізана назва' in flags:
        return "Дописати початок назви (ймовірно \"Старокостянтинівськ...\")"
    parts = []
    if 'адреса' in flags: parts.append("виправити адресу")
    if 'телефон' in flags: parts.append("виправити телефон")
    return " і ".join(parts).capitalize() if parts else "Перевірити"

for i, rec in enumerate(sorted(flagged, key=lambda r: ('артефакт (не заклад)' not in r['flags'], r['name'])), start=2):
    is_artifact = 'артефакт (не заклад)' in rec['flags']
    row = [
        rec['name'], rec['category'], ', '.join(rec['flags']), recommendation(rec['flags']),
        rec['address'], rec['phone'],
        "", "", "",
        rec['description'], rec['raw_blob']
    ]
    for col, val in enumerate(row, start=1):
        c = ws.cell(row=i, column=col, value=val)
        c.font = Font(name="Arial", size=10)
        c.border = border
        c.alignment = Alignment(vertical="top", wrap_text=True)
        if col in (7, 8, 9):
            c.fill = fill_delete if is_artifact else fill_input

widths = [22, 20, 20, 26, 24, 20, 20, 24, 18, 28, 50]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 34

wb.save('/mnt/user-data/outputs/starokostiantyniv_na_perevirku_v2.xlsx')
print('Збережено, рядків:', len(flagged))
